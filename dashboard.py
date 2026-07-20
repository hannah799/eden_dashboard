import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from collections import defaultdict
import openpyxl
import os, csv

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Eden Cannabis | Dashboard",
    page_icon="🌿",
    layout="wide",
)

# ── Styling ───────────────────────────────────────────────────────────────────
st.markdown("""
<style>
  [data-testid="stAppViewContainer"] { background: #f0f6f0; }
  [data-testid="stHeader"] { background: rgba(240,246,240,0.9); }
  .main .block-container { padding-top: 1.5rem; max-width: 1400px; margin: 0 auto; }
  #MainMenu { display: none; }
  [data-testid="stFooter"] { display: none; }

  .kpi-card {
    background: #ffffff;
    border-radius: 14px;
    padding: 20px 16px;
    box-shadow: 0 1px 5px rgba(0,0,0,0.08);
    text-align: center;
  }
  .kpi-hero {
    background: #1a4731;
    border-radius: 14px;
    padding: 20px 16px;
    box-shadow: 0 4px 14px rgba(26,71,49,0.22);
    text-align: center;
  }
  .kpi-label      { color: #6b7280; font-size: 0.69rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.09em; margin-bottom: 10px; }
  .kpi-label-hero { color: rgba(255,255,255,0.58); font-size: 0.69rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.09em; margin-bottom: 10px; }
  .kpi-value      { color: #111827; font-size: 1.8rem; font-weight: 700; line-height: 1; }
  .kpi-value-hero { color: #ffffff; font-size: 1.8rem; font-weight: 700; line-height: 1; }
  .kpi-sub        { color: #9ca3af; font-size: 0.69rem; margin-top: 7px; }
  .kpi-sub-hero   { color: rgba(255,255,255,0.46); font-size: 0.69rem; margin-top: 7px; }

  .alert-box {
    background: #d1fae5;
    border-left: 3px solid #059669;
    border-radius: 10px;
    padding: 12px 16px;
    margin-bottom: 18px;
  }
  .alert-title { color: #064e3b; font-weight: 700; font-size: 0.82rem; }
  .alert-body  { color: #065f46; font-size: 0.79rem; margin-top: 3px; }

  .story-banner {
    background: #ffffff;
    border-radius: 14px;
    padding: 16px 22px;
    box-shadow: 0 1px 5px rgba(0,0,0,0.08);
    margin-bottom: 16px;
    border-left: 4px solid #1a4731;
  }
  .story-head { color: #111827; font-size: 1.05rem; font-weight: 700; margin: 0 0 4px 0; }
  .story-sub  { color: #6b7280; font-size: 0.80rem; margin: 0; }

  div[data-testid="stRadio"] { border-bottom: 2px solid #d1d5db; margin-bottom: 24px; }
  div[data-testid="stRadio"] > div { gap: 0 !important; flex-wrap: nowrap; }
  div[data-testid="stRadio"] label {
    display: inline-flex !important; align-items: center !important;
    padding: 8px 28px 12px 28px !important;
    border-bottom: 3px solid transparent !important;
    margin-bottom: -2px !important;
    cursor: pointer !important;
    font-weight: 600 !important; font-size: 0.95rem !important;
    color: #111827 !important;
  }
  div[data-testid="stRadio"] label:has(input:checked) {
    color: #1a4731 !important;
    border-bottom: 3px solid #1a4731 !important;
  }
  div[data-testid="stRadio"] label p { color: #111827 !important; font-weight: 600 !important; }
  div[data-testid="stRadio"] label:has(input:checked) p { color: #1a4731 !important; }
  div[data-testid="stRadio"] label > div:first-child {
    width: 0 !important; height: 0 !important; margin: 0 !important;
    overflow: hidden !important; border: none !important; background: none !important;
  }

  h1 { color: #111827 !important; }
  h2, h3 { color: #1f2937 !important; }
  hr { border-color: #e5e7eb !important; }
  .data-card {
    background: #ffffff;
    border-radius: 14px;
    padding: 20px;
    box-shadow: 0 1px 5px rgba(0,0,0,0.08);
    height: 318px;
    display: flex;
    flex-direction: column;
    justify-content: space-between;
  }
  .data-table { width: 100%; border-collapse: collapse; font-size: 0.83rem; }
  .data-table th {
    color: #6b7280; font-size: 0.68rem; font-weight: 600;
    text-transform: uppercase; letter-spacing: 0.06em;
    padding: 0 10px 10px 0; text-align: left;
    border-bottom: 1px solid #e5e7eb;
  }
  .data-table td {
    color: #374151; padding: 8px 10px 8px 0;
    border-bottom: 1px solid #f3f4f6;
  }
  .data-table tr:last-child td { border-bottom: none; }
  .data-note { color: #9ca3af; font-size: 0.68rem; margin: 10px 0 0 0; }

  /* ── Filter controls ───────────────────────────────────────────────────── */
  .filter-bar {
    color: #1a4731; font-size: 0.72rem; font-weight: 700;
    text-transform: uppercase; letter-spacing: 0.09em;
    display: flex; align-items: center; gap: 8px;
    margin: 2px 0 10px 2px;
  }
  .filter-bar::before {
    content: ""; display: inline-block; width: 14px; height: 14px;
    background: #1a4731;
    -webkit-mask: url("data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'><path d='M3 4h18l-7 8v6l-4 2v-8z'/></svg>") center/contain no-repeat;
            mask: url("data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'><path d='M3 4h18l-7 8v6l-4 2v-8z'/></svg>") center/contain no-repeat;
  }
  div[data-testid="stMultiSelect"] label p,
  div[data-testid="stDateInput"]  label p {
    color: #4b5563 !important; font-size: 0.66rem !important; font-weight: 700 !important;
    text-transform: uppercase !important; letter-spacing: 0.07em !important;
  }
  div[data-testid="stMultiSelect"] div[data-baseweb="select"] > div,
  div[data-testid="stDateInput"]  div[data-baseweb="input"] {
    background: #ffffff !important;
    border: 1px solid #d8dee6 !important;
    border-radius: 10px !important;
    box-shadow: 0 1px 3px rgba(0,0,0,0.04) !important;
    min-height: 42px;
  }
  /* cap the pills area height so a long multi-select never balloons the row */
  div[data-testid="stMultiSelect"] div[data-baseweb="select"] > div:first-child {
    max-height: 78px; overflow-y: auto;
  }
  div[data-testid="stMultiSelect"] span[data-baseweb="tag"] {
    background: #e3efe8 !important; border-radius: 6px !important;
  }
  div[data-testid="stMultiSelect"] span[data-baseweb="tag"] span { color: #1a4731 !important; }
  div[data-testid="stMultiSelect"] span[data-baseweb="tag"] svg  { fill:  #1a4731 !important; }
  div[data-testid="stMultiSelect"] div[data-baseweb="select"] svg,
  div[data-testid="stDateInput"]  svg { fill: #9ca3af !important; }
  div[data-testid="stMultiSelect"] input::placeholder { color: #9ca3af !important; }
</style>
""", unsafe_allow_html=True)

# ── Color palette ─────────────────────────────────────────────────────────────
COLORS = {
    "Flower":        "#2d6a4f",
    "Concentrates":  "#2563eb",
    "Carts":         "#7c3aed",
    "Edibles":       "#d97706",
    "Pre-rolls":     "#059669",
    "Tinctures":     "#0891b2",
    "Topical":       "#db2777",
}
ROOM_COLORS = {
    "F1-Trailer": "#2d6a4f",
    "F2-Trailer": "#2563eb",
    "F3-Trailer": "#d97706",
    "F4-Trailer": "#7c3aed",
}
CHART_BG   = "#ffffff"
GRID_COLOR = "#f3f4f6"
TEXT_COLOR = "#6b7280"

G_PER_LB = 453.592


def chart_layout(fig, title="", height=340):
    fig.update_layout(
        title=dict(text=title, font=dict(color="#111827", size=13), x=0.01, xanchor="left"),
        paper_bgcolor=CHART_BG, plot_bgcolor=CHART_BG,
        font=dict(color=TEXT_COLOR, family="sans-serif"),
        height=height, margin=dict(l=10, r=10, t=40 if title else 10, b=10),
        legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color="#374151", size=11)),
        xaxis=dict(gridcolor=GRID_COLOR, zeroline=False, tickfont=dict(color=TEXT_COLOR), linecolor="#e5e7eb"),
        yaxis=dict(gridcolor=GRID_COLOR, zeroline=False, tickfont=dict(color=TEXT_COLOR), linecolor="#e5e7eb"),
    )
    return fig


# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_data(ttl=300)
def load_data():
    wb = openpyxl.load_workbook("eden_supply_reports.xlsx", read_only=True, data_only=True)

    # ── 1. Sweed Inventory (AdvancedReport) ───────────────────────────────────
    ws = wb["AdvancedReport"]
    adv_rows = list(ws.iter_rows(values_only=True))
    adv_data = [r for r in adv_rows[11:] if r[0] is not None]
    inv_records = []
    for r in adv_data:
        inv_records.append({
            "product":  r[0], "category": r[1] or "Other", "subcat": r[2], "type": r[3],
            "size": r[6], "price": float(r[7] or 0), "brand": r[8], "quality": r[9],
            "expiry": r[10], "cbd_pct": r[11], "thc_pct": r[12],
            "location": r[13], "org": r[15] or "Unknown", "qty": int(r[16] or 0),
        })
    inv_df = pd.DataFrame(inv_records)

    # ── 2. Biotrack Inventory ─────────────────────────────────────────────────
    ws2 = wb["Biotrack inventory"]
    bio_rows = list(ws2.iter_rows(values_only=True))
    bio_records = []
    for r in bio_rows[1:]:
        bio_records.append({
            "location": r[0], "room": r[1], "product": r[3], "inv_id": r[4],
            "count": float(r[5] or 0), "category": r[6], "strain": r[7], "type": r[8],
        })
    bio_df = pd.DataFrame(bio_records)

    # ── 3. Sweed Sales ────────────────────────────────────────────────────────
    ws3 = wb["Sweed-Sales YTD"]
    sw_rows = list(ws3.iter_rows(values_only=True))
    sw_records = []
    for r in sw_rows[1:]:
        if not r[1]: continue
        sw_records.append({
            "store": r[0] or "Unknown", "ts": r[1], "category": r[2] or "Other",
            "gross": float(r[4] or 0), "net": float(r[5] or 0), "qty": float(r[6] or 0),
        })
    sw_df = pd.DataFrame(sw_records)
    if not sw_df.empty:
        sw_df["date"] = pd.to_datetime(sw_df["ts"])
        sw_df["week"]  = sw_df["date"].dt.to_period("W").apply(lambda x: x.start_time)
        sw_df["month"] = sw_df["date"].dt.to_period("M").apply(lambda x: x.start_time)

    # ── 4. Alleaves Sales ─────────────────────────────────────────────────────
    ws4 = wb["Alleaves Sales -YTD"]
    al_rows = list(ws4.iter_rows(values_only=True))
    al_records = []
    for r in al_rows[2:]:
        if not r[3]: continue
        al_records.append({
            "location": r[0] or "Unknown", "category": (r[1] or "").strip(),
            "product": r[3], "qty": float(r[4] or 0),
            "gross": float(r[6] or 0), "net": float(r[7] or 0),
        })
    al_df = pd.DataFrame(al_records)

    return inv_df, bio_df, sw_df, al_df


@st.cache_data(ttl=300)
def load_store_sales():
    """Load per-store exports for all 5 locations — Tampa, Sarasota, Cocoa Beach,
    Orlando, Delivery Hub (May 1–Jun 1 2026 snapshot)."""
    import os

    def _cat_from_product(name):
        n = (name or "").lower()
        if "whole flower" in n:                          return "Flower"
        if "pre-roll" in n:                              return "Pre-rolls"
        if "vape pen" in n or "disposable" in n:        return "Carts"
        if "rosin" in n:                                 return "Concentrates"
        if "edible" in n:                                return "Edibles"
        if "tincture" in n:                              return "Tinctures"
        if "topical" in n:                               return "Topical"
        return "Other"

    store_map = {}
    for f in os.listdir("."):
        if not f.endswith(".xlsx") or not f.startswith("Sales"):
            continue
        if "COCOA BEACH" in f:                          store_map[f] = "Eden - Cocoa Beach"
        elif "ORLANDO" in f:                            store_map[f] = "Eden - Orlando"
        elif "SARASOTA" in f and "DELIVERY" not in f:  store_map[f] = "Eden - Sarasota"
        elif "DELIVERY" in f:                           store_map[f] = "Eden - Delivery Hub"
        elif "TAMPA" in f:                              store_map[f] = "Eden - Tampa"

    records = []
    for fname, store in store_map.items():
        wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # rows[0-2]=metadata, rows[3]=header, rows[4]=TOTAL (skip), rows[5+]=products
        # cols: Product[0] SKU[1] QTY[2] Sales Receipts[3] Gross Sales[4] Discounts[5]
        for r in rows[5:]:
            if not r[0] or r[0] == "TOTAL":
                continue
            records.append({
                "store":    store,
                "product":  r[0],
                "category": _cat_from_product(r[0]),
                "qty":      float(r[2] or 0),
                "gross":    float(r[4] or 0),
                "net":      float(r[3] or 0),   # Sales Receipts = gross − discounts − rewards
            })
        wb.close()

    if records:
        return pd.DataFrame(records)
    return pd.DataFrame(columns=["store", "product", "category", "qty", "gross", "net"])


inv_df, bio_df, sw_df, al_df = load_data()
store_df = load_store_sales()


# ── Monday.com board exports (monday_reports_july_19/) ─────────────────────────
MONDAY_DIR = "monday_reports_july_19"

def _parse_monday(path):
    """Monday.com board export → list of dicts keyed by column header.
    Skips the board title/description rows, repeated 'Name' header rows,
    'Subitems' sub-table headers, and group-title rows (single populated cell).
    Each returned row carries `_group` = the most recent group title."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    header = None
    group = None
    out = []
    for r in rows:
        if not r or all(c is None for c in r):
            continue
        if r[0] in ("Name", "Subitems"):        # header / subitem-header row
            if r[0] == "Name":
                header = [str(h) for h in r]
            continue
        nonnull = [c for c in r if c is not None]
        if len(nonnull) == 1 and r[0] is not None:   # group title
            group = str(r[0])
            continue
        if header is None:
            continue
        d = {header[i]: (r[i] if i < len(r) else None) for i in range(len(header))}
        d["_group"] = group
        out.append(d)
    return out

def _mnum(x):
    return float(x) if isinstance(x, (int, float)) else None


@st.cache_data(ttl=300)
def load_harvest():
    # Sourced from the Monday "Harvest Tracker" board export. Rows are strain-
    # level within a cycle; the blank-name cycle-summary rows are dropped so the
    # per-strain weights (which is what every downstream chart expects) are used.
    raw = _parse_monday(os.path.join(MONDAY_DIR, "Harvest_Tracker_1784506120.xlsx"))
    h_records = []
    for d in raw:
        name = d.get("Name")
        if not name or str(name).strip() == "":
            continue
        hd = d.get("Date of Harvest")
        td = d.get("Test Date")
        h_records.append({
            "strain":           name,
            "harvest_date":     hd if hasattr(hd, "year") else None,
            "status":           d.get("Status"),
            "cycle":            d.get("Cycle #"),
            "room":             d.get("Flower Room"),
            "num_plants":       _mnum(d.get("# of Plants")),
            "fresh_frozen_g":   _mnum(d.get("Fresh Frozen Weight")),
            "wet_weight_g":     _mnum(d.get("Wet Weight")),
            "dry_weight_g":     _mnum(d.get("Dry Weight Actual")),
            "sf_canopy":        _mnum(d.get("SF")),
            "grams_per_sf":     _mnum(d.get("Gr/ SF Actual")),
            "trim_weight_g":    _mnum(d.get("Trim Weight")),
            "trimmed_flower_g": _mnum(d.get("Trimmed Flower")),
            "est_35_units":     _mnum(d.get("Est. 3.5g units")),
            "est_1g_units":     _mnum(d.get("Est. 1g units")),
            "designation":      d.get("Designation"),
            "thc_pct":          _mnum(d.get("THC%")),
            "terps_pct":        _mnum(d.get("TOTAL TERPS%")),
            "test_date":        td if hasattr(td, "year") else None,
            "grade":            d.get("Grade"),
        })
    return pd.DataFrame(h_records)


@st.cache_data(ttl=300)
def load_monday_inventory():
    """Production supplies (reorder status) + finished-goods batch tracker."""
    out = {}

    # ── Production supplies ────────────────────────────────────────────────────
    sup_raw = _parse_monday(os.path.join(MONDAY_DIR,
                                          "Inventory_Management_Production_1784506111.xlsx"))
    sup = []
    for d in sup_raw:
        if not d.get("Name") or d.get("Status") not in (
                "Reorder Needed", "Emergency Order Needed", "Sufficient"):
            continue
        sup.append({
            "item": d.get("Name"), "unit": d.get("Unit Type"),
            "on_hand": _mnum(d.get("Count on Hand")),
            "reorder_level": _mnum(d.get("Reorder Level")),
            "max_level": _mnum(d.get("Max Level")),
            "status": d.get("Status"), "vendor": d.get("Regular Vendor"),
        })
    out["supplies"] = pd.DataFrame(sup)

    # ── Finished goods batches ─────────────────────────────────────────────────
    fg_raw = _parse_monday(os.path.join(MONDAY_DIR,
                                        "Finished_Goods_Tracker_1784506133.xlsx"))
    fg = []
    for d in fg_raw:
        if not d.get("Name"):
            continue
        exp = d.get("Expiration Date")
        fg.append({
            "product": d.get("Name"), "group": d.get("_group"),
            "category": d.get("Product Category"),
            "strain": d.get("Strain/Flavor"), "size": d.get("Size"),
            "on_hand": _mnum(d.get("QTY - ON HAND")) or 0,
            "product_status": d.get("Product Status"),
            "pass_fail": d.get("Pass/Fail"),
            "thc_pct": _mnum(d.get("THC%")),
            "days_to_exp": _mnum(d.get("Days until Expiration")),
            "exp_date": exp if hasattr(exp, "year") else None,
        })
    out["finished_goods"] = pd.DataFrame(fg)

    return out

_monday_inv_error = None
try:
    MINV = load_monday_inventory()
except Exception as e:
    _monday_inv_error = str(e)
    MINV = {}


# ══════════════════════════════════════════════════════════════════════════════
# HEADSET ANALYTICS (headset_reports_july_19/) — weekly sales, top products,
# days-supply, dead stock, purchase-method (delivery vs in-store), demographics.
# ══════════════════════════════════════════════════════════════════════════════
HEADSET_DIR = "headset_reports_july_19"

def _hs_money(x):
    if x is None:
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace("$", "").replace(",", "").replace("%", "")
    if s in ("", "-", "‎"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0

def _hs_clean_store(s):
    return str(s).replace("Eden FL - ", "").replace("Eden - ", "").strip()

@st.cache_data(ttl=300)
def load_headset():
    def _rows(fname):
        wb = openpyxl.load_workbook(os.path.join(HEADSET_DIR, fname),
                                    read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        return rows

    def _read_csv(fname):
        with open(os.path.join(HEADSET_DIR, fname), newline="", encoding="utf-8-sig") as f:
            return list(csv.reader(f))

    out = {}

    # 1. Weekly revenue by store (Mar–Jul 2026)
    rows = _rows("Sales from Last 26 complete Weeks.xlsx")
    stores = [_hs_clean_store(c) for c in rows[0][1:5]]
    recs = []
    for r in rows[2:]:
        if not r[0]:
            continue
        wk = pd.to_datetime(r[0])
        for i, st_name in enumerate(stores):
            v = r[1 + i]
            if v is not None:
                recs.append({"week": wk, "store": st_name, "gross": _hs_money(v)})
    out["weekly"] = pd.DataFrame(recs)

    # 2. Top products (with gross margin)
    rows = _rows("Top Products.xlsx")
    recs = []
    for r in rows[1:]:
        if not r[1]:
            continue
        recs.append({
            "product": r[1], "category": r[2], "brand": r[4],
            "gross": _hs_money(r[5]), "units": _hs_money(r[6]),
            "gm_pct": (float(r[8]) * 100 if r[8] is not None else None),
            "avg_price": _hs_money(r[9]),
        })
    out["top_products"] = pd.DataFrame(recs)

    # 3. Days of supply by category
    rows = _rows("Days Supply.xlsx")
    recs = []
    for r in rows[1:]:
        if not r[0]:
            continue
        recs.append({"category": r[0], "on_hand": _hs_money(r[1]),
                     "per_day": _hs_money(r[2]), "days_remaining": _hs_money(r[3])})
    out["days_supply"] = pd.DataFrame(recs)

    # 4. Dead stock — SKUs in stock with no sales
    rows = _rows("Products In-Stock with No Sales.xlsx")
    recs = []
    for r in rows[1:]:
        if not r[1]:
            continue
        recs.append({"category": r[1], "skus": int(_hs_money(r[2]))})
    out["dead_stock"] = pd.DataFrame(recs)

    # 5. Product list → reorder watchlist
    rows = _rows("Product List (1).xlsx")
    recs = []
    for r in rows[1:]:
        if not r[2]:
            continue
        recs.append({
            "store": _hs_clean_store(r[1]), "product": r[2], "category": r[5],
            "on_hand": _hs_money(r[7]), "sold": _hs_money(r[9]),
            "days_remaining": (None if r[10] is None else _hs_money(r[10])),
            "per_day": (None if r[11] is None else _hs_money(r[11])),
            "suggested_order": _hs_money(r[18]),
        })
    out["product_list"] = pd.DataFrame(recs)

    # 6. Purchase-method summary (delivery vs in-store KPIs)
    c = _read_csv("pm_summary.csv")
    recs = []
    for r in c[1:]:
        if len(r) < 6 or not r[1]:
            continue
        recs.append({"flag": r[1], "gross": _hs_money(r[2]), "avg_basket": _hs_money(r[3]),
                     "units_per_basket": _hs_money(r[4]), "discount_pct": _hs_money(r[5])})
    out["pm_summary"] = pd.DataFrame(recs)

    # 7. Purchase-method weekly revenue
    c = _read_csv("pm_sales_performance.csv")
    recs = []
    for r in c[2:]:
        if not r or not r[0]:
            continue
        recs.append({"week": pd.to_datetime(r[0]),
                     "Delivery": _hs_money(r[1]), "In-Store": _hs_money(r[2])})
    out["pm_weekly"] = pd.DataFrame(recs)

    # 8. Demographics — revenue by age group × gender (summed over delivery flag)
    c = _read_csv("pm_demographics.csv")
    recs = []
    for r in c[1:]:
        if len(r) < 4 or not r[0]:
            continue
        recs.append({"age": r[0], "gender": r[1], "gross": _hs_money(r[3])})
    demo = pd.DataFrame(recs)
    if not demo.empty:
        demo = demo.groupby(["age", "gender"], as_index=False)["gross"].sum()
    out["demographics"] = demo

    # 9. Sales by weekday × hour (summed over delivery flag)
    c = _read_csv("pm_weekday_hour.csv")
    day_of_col = {i: c[0][i] for i in range(2, len(c[0]))}
    recs = []
    for r in c[3:]:
        if len(r) < 3 or not r[1]:
            continue
        try:
            hour = int(r[1])
        except ValueError:
            continue
        for i in range(2, len(r)):
            day = day_of_col.get(i)
            if not day:
                continue
            recs.append({"day": day, "hour": hour, "gross": _hs_money(r[i])})
    wh = pd.DataFrame(recs)
    if not wh.empty:
        wh = wh.groupby(["day", "hour"], as_index=False)["gross"].sum()
    out["weekday_hour"] = wh

    return out

_headset_error = None
try:
    HS = load_headset()
except Exception as e:
    _headset_error = str(e)
    HS = {}


# ══════════════════════════════════════════════════════════════════════════════
# PRE-COMPUTE AGGREGATES (supply & demand)
# ══════════════════════════════════════════════════════════════════════════════
CAT_MAP = {
    "Flower": "Flower", "Flower ": "Flower",
    "Live Rosin": "Concentrates", "Live Rosin (Disposable)": "Carts",
    "Edibles": "Edibles", "Tincture": "Tinctures", "Topical": "Topical",
}
al_df["cat_norm"] = al_df["category"].map(CAT_MAP).fillna(al_df["category"])
LOC_MAP = {
    "EDEN Tampa": "Eden - Tampa", "EDEN Sarasota": "Eden - Sarasota",
    "Tavares Delivery": "Eden - Delivery Hub",
}
al_df["loc_norm"] = al_df["location"].map(LOC_MAP).fillna(al_df["location"])

# Exclude Tampa from Alleaves — sw_df covers Tampa more completely (Mar–Jun vs Jan–Apr)
al_notampa = al_df[al_df["loc_norm"] != "Eden - Tampa"]

# Tampa's May 2026 is supplied by its per-store file (store_df). Drop May from the Sweed
# series so each Tampa month is counted exactly once in blended totals & breakdowns.
# (Full sw_df is kept for the Weeks-of-Stock velocity, which needs daily data.)
sw_dedup = sw_df[~((sw_df["date"].dt.year == 2026) & (sw_df["date"].dt.month == 5))].copy()

ytd_gross = sw_dedup["gross"].sum() + al_notampa["gross"].sum() + store_df["gross"].sum()
ytd_net   = sw_dedup["net"].sum()   + al_notampa["net"].sum()   + store_df["net"].sum()
ytd_qty   = sw_dedup["qty"].sum()   + al_notampa["qty"].sum()  + store_df["qty"].sum()
sw_total_gross = sw_df["gross"].sum()
n_locations = 5  # Tampa, Sarasota, Cocoa Beach, Orlando, Delivery Hub

EXCLUDE_LOCS = {"Quarantine", "Quarantine 1283", "temp hoLDING"}
fin_df = inv_df[~inv_df["location"].isin(EXCLUDE_LOCS)]
total_finished = fin_df["qty"].sum()

# Biotrack pipeline
PIPELINE_ROOMS = {
    "02.1 - Hash Processing Room":             "Concentrate Processing",
    "02.2 - Hash Washing Room (Freezer)":      "Concentrate Processing",
    "02.3 - Hash Processing Room (WIP)":       "Concentrate Processing",
    "02.4 - Hash Processing Room Storage Cabinet": "Concentrate Processing",
    "03 - Trimming (WIP)":                     "Trimming",
    "03.1 - Trim Room":                        "Trimming",
    "TAV-PROD-002 - Kitchen (WIP)":            "Edible Production",
    "TAV-PROD-004-Kitchen Storage":            "Edible Production",
    "04 - Ready for Packaging":                "Ready for Packaging",
    "05 - Packaging (WIP)":                    "Packaging WIP",
    "12 - CGMP Packaging Room":                "CGMP Packaging",
    "12.1 CGMP Finished Product Fridge":       "CGMP Finished",
    "12.2 CGMP WIP Fridge":                    "CGMP WIP",
    "07 - Awaiting COA":                       "Awaiting Test Results",
    "09 - Approved for Labeling":              "Approved for Labeling ✓",
    "13 - Processing Vault":                   "Processing Vault",
}
bio_df["stage"] = bio_df["room"].map(PIPELINE_ROOMS)

BIO_CAT_MAP = {
    "Whole Flower": "Flower", "Trim/Shake": "Flower",
    "Pre-roll Material": "Flower", "Pre-Rolls": "Pre-rolls",
    "Hash": "Concentrates", "Fresh Frozen Material": "Concentrates",
    "Rosin": "Concentrates", "Live Rosin": "Concentrates", "Kief": "Concentrates",
    "510 Thread": "Carts", "510-Threads": "Carts", "Rosin (Cart)": "Carts",
    "Edible": "Edibles", "Tincture": "Tinctures", "Syringe": "Concentrates",
}
bio_df["broad_cat"] = bio_df["category"].map(BIO_CAT_MAP).fillna("Other")

FLOWER_CATS_BIO = {"Whole Flower", "Trim/Shake", "Pre-roll Material", "Fresh Frozen Material"}
MFG_CATS_BIO    = {"Hash", "Live Rosin", "Rosin", "Kief"}
G_PER_L = 1000.0

def _bio_convert(row):
    cat = row["category"] or ""
    g = row["count"]
    if cat in FLOWER_CATS_BIO: return g / G_PER_LB, "lbs"
    if cat in MFG_CATS_BIO:    return g / G_PER_L,  "L"
    return g, "units"

bio_df[["display_count", "unit"]] = bio_df.apply(_bio_convert, axis=1, result_type="expand")

STAGE_ORDER = [
    "Concentrate Processing", "Edible Production", "Trimming",
    "Ready for Packaging", "Packaging WIP", "CGMP WIP", "CGMP Packaging",
    "CGMP Finished", "Awaiting Test Results", "Approved for Labeling ✓", "Processing Vault",
]

VAULT_LOCS = {"TAV-Stock Room"}
DISP_ORGS  = {
    "Eden - Tampa": "Tampa", "Eden - Sarasota": "Sarasota",
    "Eden - Cocoa Beach": "Cocoa Beach", "Eden - Orlando": "Orlando",
}
vault_df = fin_df[fin_df["location"].isin(VAULT_LOCS)]
disp_df  = fin_df[fin_df["org"].isin(DISP_ORGS.keys()) & ~fin_df["location"].isin(VAULT_LOCS)]

# Harvest pre-compute — always produces DataFrames with correct columns
_harvest_error = None
_EMPTY_H = pd.DataFrame(columns=[
    "strain","harvest_date","status","cycle","room","num_plants",
    "fresh_frozen_g","wet_weight_g","dry_weight_g","sf_canopy","grams_per_sf",
    "trim_weight_g","trimmed_flower_g","est_35_units","est_1g_units","designation",
    "thc_pct","terps_pct","test_date","grade",
    "dry_lbs","wet_lbs","ff_lbs","trim_lbs","flower_lbs","month",
])
try:
    harvest_df = load_harvest()
    h = harvest_df.copy()
    h["harvest_date"] = pd.to_datetime(h["harvest_date"], errors="coerce")
    h["dry_lbs"]    = h["dry_weight_g"]  / G_PER_LB
    h["wet_lbs"]    = h["wet_weight_g"]  / G_PER_LB
    h["ff_lbs"]     = h["fresh_frozen_g"] / G_PER_LB
    h["trim_lbs"]   = h["trim_weight_g"] / G_PER_LB
    h["flower_lbs"] = h["trimmed_flower_g"] / G_PER_LB
    h["month"] = h["harvest_date"].dt.to_period("M").apply(
        lambda x: x.start_time if pd.notna(x) else pd.NaT
    )
    harvested = h[h["status"] == "Harvested"].copy()
    active    = h[h["status"].isin(["Flowering", "Moved to Trim Room"])].copy()
    upcoming  = h[h["status"] == "Upcoming Cycle"].copy()
except Exception as e:
    _harvest_error = str(e)
    harvested = active = upcoming = _EMPTY_H.copy()

# Pipeline bulk totals (needed in narrative banner)
pipeline_bio = bio_df[
    bio_df["stage"].notna() &
    ~bio_df["stage"].str.contains("Approved|Sample", na=False)
]
flower_lbs = pipeline_bio[pipeline_bio["unit"] == "lbs"]["display_count"].sum()
mfg_L      = pipeline_bio[pipeline_bio["unit"] == "L"]["display_count"].sum()

# ── Header ───────────────────────────────────────────────────────────────────
st.title("Eden Cannabis")

# Narrative banner — all-store top category (Tampa Sweed + Alleaves + per-store files)
_bcat = pd.concat([
    sw_dedup.groupby("category")["gross"].sum(),
    al_notampa.groupby("cat_norm")["gross"].sum(),
    store_df.groupby("category")["gross"].sum(),
]).groupby(level=0).sum()
if not _bcat.empty and _bcat.sum() > 0:
    top_cat_all = _bcat.idxmax()
    top_pct_all = _bcat.max() / _bcat.sum() * 100
else:
    top_cat_all, top_pct_all = "Flower", 0
st.markdown(f"""<div class="story-banner">
  <p class="story-head">YTD: ${ytd_gross:,.0f} gross revenue &nbsp;·&nbsp; {int(ytd_qty):,} units sold across all locations</p>
  <p class="story-sub">{top_cat_all} leads all-store sales at {top_pct_all:.0f}% of revenue &nbsp;·&nbsp; {total_finished:,} finished units on-hand &nbsp;·&nbsp; {flower_lbs:,.0f} lbs flower + {mfg_L:.1f} L concentrate in production pipeline</p>
</div>""", unsafe_allow_html=True)


view = st.radio("", ["Retail Sales", "Inventory", "Harvest"], horizontal=True, label_visibility="collapsed")


# ── Filter helpers ─────────────────────────────────────────────────────────────
def _clean_store(s):
    return str(s).replace("Eden - ", "")

def _options(*series):
    vals = set()
    for s in series:
        vals |= {v for v in pd.Series(list(s)).dropna().unique()}
    return sorted(vals, key=str)


if view == "Retail Sales":

    # ══════════════════════════════════════════════════════════════════════════
    # EXECUTIVE SUMMARY — authoritative company actuals through July.
    # Transcribed from executive_reports_july_19/ (Power BI export, snapshot
    # Jul 19 2026). Monthly net sales use the KPI-panel figures (internally
    # consistent; June & July also tie out to the per-door tables). May's
    # per-door screenshot was a partial capture, so only Jun & Jul carry a
    # reliable store-level breakdown. July is month-to-date (through Jul 19).
    # ══════════════════════════════════════════════════════════════════════════
    EXEC_MONTHLY = [
        {"month": "May 2026",       "net": 55125.38,  "txns": 687, "patients": 468, "new": 377, "returning": 91,  "mtd": False},
        {"month": "Jun 2026",       "net": 100087.80, "txns": 943, "patients": 645, "new": 512, "returning": 133, "mtd": False},
        {"month": "Jul 2026 (MTD)", "net": 50743.40,  "txns": 538, "patients": 403, "new": 252, "returning": 151, "mtd": True},
    ]
    EXEC_DOORS = {   # gross, net, unique patients per store
        "Jun 2026": [
            ("Cocoa Beach", 56088.0, 37740.73, 271), ("Orlando", 36852.5, 25259.17, 157),
            ("Sarasota", 29952.0, 20063.16, 130), ("Tampa", 21816.5, 15249.66, 87),
            ("Delivery Hub", 2780.0, 1775.08, 7),
        ],
        "Jul 2026 (MTD)": [
            ("Cocoa Beach", 23110.0, 15815.58, 158), ("Sarasota", 19707.5, 12966.65, 91),
            ("Orlando", 17477.0, 11865.85, 95), ("Tampa", 9638.5, 6793.34, 53),
            ("Delivery Hub", 4967.5, 3303.05, 9),
        ],
    }

    st.subheader("Executive Summary")
    st.caption("Company actuals through July · executive reports (Power BI), snapshot Jul 19, 2026 · July is month-to-date")

    _jul = EXEC_MONTHLY[-1]
    _jun = EXEC_MONTHLY[-2]
    e1, e2, e3, e4, e5 = st.columns(5)
    with e1:
        st.markdown(f"""<div class="kpi-hero">
          <div class="kpi-label-hero">Net Sales · July MTD</div>
          <div class="kpi-value-hero">${_jul['net']:,.0f}</div>
          <div class="kpi-sub-hero">June (full month): ${_jun['net']:,.0f}</div>
        </div>""", unsafe_allow_html=True)
    for col, label, value, sub in [
        (e2, "Patients · July",      f"{_jul['patients']:,}",  f"{_jun['patients']:,} in June"),
        (e3, "New Patients · July",  f"{_jul['new']:,}",       f"{_jul['returning']:,} returning"),
        (e4, "Transactions · July",  f"{_jul['txns']:,}",      f"{_jun['txns']:,} in June"),
        (e5, "Avg Ticket · July",    f"${_jul['net']/_jul['txns']:,.0f}", "Net ÷ transactions"),
    ]:
        with col:
            st.markdown(f"""<div class="kpi-card">
              <div class="kpi-label">{label}</div>
              <div class="kpi-value">{value}</div>
              <div class="kpi-sub">{sub}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<div style='height:14px;'></div>", unsafe_allow_html=True)
    col_ns, col_pat = st.columns(2)
    with col_ns:
        _labels = [m["month"] for m in EXEC_MONTHLY]
        _nets   = [m["net"] for m in EXEC_MONTHLY]
        _bar_c  = ["#74c69d" if m["mtd"] else "#1a4731" for m in EXEC_MONTHLY]
        fig_ns = go.Figure(go.Bar(
            x=_labels, y=_nets, marker_color=_bar_c,
            text=[f"${v:,.0f}" for v in _nets], textposition="outside", cliponaxis=False,
            hovertemplate="%{x}<br>$%{y:,.0f} net sales<extra></extra>",
        ))
        chart_layout(fig_ns, "Company Net Sales by Month", height=300)
        fig_ns.update_layout(yaxis=dict(range=[0, max(_nets) * 1.2]), margin=dict(l=10, r=10, t=40, b=10))
        st.plotly_chart(fig_ns, use_container_width=True)

    with col_pat:
        fig_pat = go.Figure()
        fig_pat.add_trace(go.Bar(
            name="New", x=[m["month"] for m in EXEC_MONTHLY], y=[m["new"] for m in EXEC_MONTHLY],
            marker_color="#2d6a4f", hovertemplate="%{x}<br>%{y} new<extra></extra>"))
        fig_pat.add_trace(go.Bar(
            name="Returning", x=[m["month"] for m in EXEC_MONTHLY], y=[m["returning"] for m in EXEC_MONTHLY],
            marker_color="#d97706", hovertemplate="%{x}<br>%{y} returning<extra></extra>"))
        fig_pat.update_layout(barmode="stack", legend=dict(orientation="h", y=1.15), yaxis_title="")
        chart_layout(fig_pat, "Unique Patients — New vs Returning", height=300)
        st.plotly_chart(fig_pat, use_container_width=True)

    # Store-level door breakdown — June (last full month) vs July MTD
    door_rows = ["Cocoa Beach", "Sarasota", "Orlando", "Tampa", "Delivery Hub"]
    jun_map = {d[0]: d for d in EXEC_DOORS["Jun 2026"]}
    jul_map = {d[0]: d for d in EXEC_DOORS["Jul 2026 (MTD)"]}
    rows_html = ""
    for store in door_rows:
        j = jun_map.get(store); k = jul_map.get(store)
        rows_html += (
            f"<tr><td>{store}</td>"
            f"<td style='text-align:right;'>${j[1]:,.0f}</td>"
            f"<td style='text-align:right;'>${j[2]:,.0f}</td>"
            f"<td style='text-align:right;'>{j[3]}</td>"
            f"<td style='text-align:right;border-left:1px solid #e5e7eb;'>${k[1]:,.0f}</td>"
            f"<td style='text-align:right;'>${k[2]:,.0f}</td>"
            f"<td style='text-align:right;'>{k[3]}</td></tr>"
        )
    st.markdown(f"""<div class="data-card" style="height:auto;">
      <table class="data-table">
        <thead>
          <tr><th></th><th colspan="3" style="text-align:center;border-bottom:1px solid #e5e7eb;">June (full month)</th>
              <th colspan="3" style="text-align:center;border-left:1px solid #e5e7eb;border-bottom:1px solid #e5e7eb;">July (MTD)</th></tr>
          <tr><th>Store</th>
              <th style="text-align:right;">Gross</th><th style="text-align:right;">Net</th><th style="text-align:right;">Patients</th>
              <th style="text-align:right;border-left:1px solid #e5e7eb;">Gross</th><th style="text-align:right;">Net</th><th style="text-align:right;">Patients</th></tr>
        </thead>
        <tbody>{rows_html}</tbody>
      </table>
      <p class="data-note">Net sales per door · executive reports. July is month-to-date through Jul 19.</p>
    </div>""", unsafe_allow_html=True)

    st.divider()

    # ── Filters ───────────────────────────────────────────────────────────────
    _cat_opts   = _options(sw_dedup["category"], al_notampa["cat_norm"], store_df["category"])
    _store_opts = _options(
        sw_dedup["store"].map(_clean_store),
        al_notampa["loc_norm"].map(_clean_store),
        store_df["store"].map(_clean_store),
    )
    _dmin, _dmax = sw_dedup["date"].min().date(), sw_dedup["date"].max().date()
    st.markdown('<div class="filter-bar">Filters</div>', unsafe_allow_html=True)
    fc1, fc2, fc3 = st.columns([2, 2, 2])
    sel_range  = fc1.date_input("Date range (Tampa daily sales)", (_dmin, _dmax),
                                min_value=_dmin, max_value=_dmax, key="rs_date")
    sel_cats   = fc2.multiselect("Category", _cat_opts, default=[],
                                 placeholder="All categories", key="rs_cat") or _cat_opts
    sel_stores = fc3.multiselect("Store", _store_opts, default=[],
                                 placeholder="All stores", key="rs_store") or _store_opts

    if isinstance(sel_range, (list, tuple)) and len(sel_range) == 2:
        d0, d1 = pd.Timestamp(sel_range[0]), pd.Timestamp(sel_range[1]) + pd.Timedelta(days=1)
    else:
        d0, d1 = pd.Timestamp(_dmin), pd.Timestamp(_dmax) + pd.Timedelta(days=1)

    # Date filter applies only to dated Tampa sales; Alleaves & per-store files are
    # single snapshots with no per-row date, so only category/store filter them.
    sw_dedup = sw_dedup[
        (sw_dedup["date"] >= d0) & (sw_dedup["date"] < d1) &
        sw_dedup["category"].isin(sel_cats) &
        sw_dedup["store"].map(_clean_store).isin(sel_stores)
    ].copy()
    al_notampa = al_notampa[
        al_notampa["cat_norm"].isin(sel_cats) &
        al_notampa["loc_norm"].map(_clean_store).isin(sel_stores)
    ].copy()
    store_df = store_df[
        store_df["category"].isin(sel_cats) &
        store_df["store"].map(_clean_store).isin(sel_stores)
    ].copy()

    # Recompute headline KPIs from the filtered frames
    ytd_gross   = sw_dedup["gross"].sum() + al_notampa["gross"].sum() + store_df["gross"].sum()
    ytd_net     = sw_dedup["net"].sum()   + al_notampa["net"].sum()   + store_df["net"].sum()
    ytd_qty     = sw_dedup["qty"].sum()   + al_notampa["qty"].sum()   + store_df["qty"].sum()
    n_locations = len(sel_stores)
    _date_sub   = f"{d0:%b %-d} – {(d1 - pd.Timedelta(days=1)):%b %-d, %Y}"
    _loc_sub    = ", ".join(sel_stores) if len(sel_stores) <= 3 else f"{n_locations} locations"

    # ── Sales KPIs ────────────────────────────────────────────────────────────
    # Combined gross by category across all sources (Tampa Sweed + Alleaves + 5 store files)
    _sw_c    = sw_dedup.groupby("category").agg(gross=("gross", "sum")).reset_index()
    _al_c    = al_notampa.groupby("cat_norm").agg(gross=("gross", "sum")).reset_index().rename(columns={"cat_norm": "category"})
    _store_c = store_df.groupby("category").agg(gross=("gross", "sum")).reset_index()
    cat_all  = pd.concat([_sw_c, _al_c, _store_c]).groupby("category").agg(gross=("gross", "sum")).reset_index()
    if not cat_all.empty:
        _top      = cat_all.sort_values("gross", ascending=False).iloc[0]
        top_cat   = _top["category"]
        top_share = _top["gross"] / cat_all["gross"].sum() * 100
    else:
        top_cat, top_share = "–", 0

    k1, k2, k3, k4, k5 = st.columns(5)
    with k1:
        st.markdown(f"""<div class="kpi-hero">
          <div class="kpi-label-hero">YTD Gross Revenue</div>
          <div class="kpi-value-hero">${ytd_gross:,.0f}</div>
          <div class="kpi-sub-hero">{_date_sub}</div>
        </div>""", unsafe_allow_html=True)
    for col, label, value, sub in [
        (k2, "Net Revenue",      f"${ytd_net:,.0f}",          "After discounts"),
        (k3, "Units Sold",       f"{int(ytd_qty):,}",         "Selected filters"),
        (k4, "Locations",        f"{n_locations}",            _loc_sub),
        (k5, "Top Category",     f"{top_cat}",                f"{top_share:.0f}% of gross sales"),
    ]:
        with col:
            st.markdown(f"""<div class="kpi-card">
              <div class="kpi-label">{label}</div>
              <div class="kpi-value">{value}</div>
              <div class="kpi-sub">{sub}</div>
            </div>""", unsafe_allow_html=True)

    st.divider()

    # ── Sales ─────────────────────────────────────────────────────────────────
    st.subheader("YTD Sales")

    col_cat, col_loc = st.columns(2)
    with col_cat:
        sw_cat    = sw_dedup.groupby("category").agg(gross=("gross","sum")).reset_index()
        al_cat    = al_notampa.groupby("cat_norm").agg(gross=("gross","sum")).reset_index()
        al_cat.rename(columns={"cat_norm": "category"}, inplace=True)
        store_cat = store_df.groupby("category").agg(gross=("gross","sum")).reset_index()
        cat_total = pd.concat([sw_cat, al_cat, store_cat]).groupby("category").agg(
            gross=("gross","sum")).reset_index().sort_values("gross", ascending=True)
        fig5 = go.Figure(go.Bar(
            x=cat_total["gross"], y=cat_total["category"], orientation="h",
            marker_color=[COLORS.get(c, "#2563eb") for c in cat_total["category"]],
            text=cat_total["gross"].apply(lambda x: f"${x:,.0f}"),
            textposition="outside", textfont=dict(color="#374151", size=11),
            cliponaxis=False,
            hovertemplate="%{y}  $%{x:,.0f}<extra></extra>",
        ))
        chart_layout(fig5, "Gross Sales by Category", height=300)
        fig5.update_layout(
            xaxis_title="", xaxis=dict(range=[0, cat_total["gross"].max() * 1.25]),
            yaxis=dict(gridcolor="rgba(0,0,0,0)"),
            margin=dict(l=10, r=80, t=40, b=10),
        )
        st.plotly_chart(fig5, use_container_width=True)

    with col_loc:
        sw_loc    = sw_dedup.groupby("store").agg(gross=("gross","sum")).reset_index()
        sw_loc.rename(columns={"store": "location"}, inplace=True)
        al_loc    = al_notampa.groupby("loc_norm").agg(gross=("gross","sum")).reset_index()
        al_loc.rename(columns={"loc_norm": "location"}, inplace=True)
        store_loc = store_df.groupby("store").agg(gross=("gross","sum")).reset_index()
        store_loc.rename(columns={"store": "location"}, inplace=True)
        loc_combined = (
            pd.concat([sw_loc, al_loc, store_loc])
            .groupby("location").agg(gross=("gross","sum"))
            .reset_index()
            .sort_values("gross", ascending=True)
        )
        loc_combined["loc_short"] = loc_combined["location"].str.replace("Eden - ", "", regex=False)
        fig6 = go.Figure(go.Bar(
            x=loc_combined["gross"], y=loc_combined["loc_short"], orientation="h",
            marker_color="#2d6a4f",
            text=loc_combined["gross"].apply(lambda x: f"${x:,.0f}"),
            textposition="outside", textfont=dict(color="#374151", size=11),
            cliponaxis=False,
            hovertemplate="%{y}  $%{x:,.0f}<extra></extra>",
        ))
        chart_layout(fig6, "Gross Sales by Location", height=300)
        fig6.update_layout(
            xaxis_title="",
            xaxis=dict(range=[0, loc_combined["gross"].max() * 1.30]),
            yaxis=dict(gridcolor="rgba(0,0,0,0)"),
            margin=dict(l=10, r=110, t=40, b=10),
        )
        st.plotly_chart(fig6, use_container_width=True)

    col_trend, col_pie = st.columns([3, 2])
    with col_trend:
        # Tampa: monthly totals from Sweed, May excluded (dedup) — Mar, Apr, Jun.
        # Tampa's May arrives below via its per-store file, so it is counted once.
        tampa_monthly = (
            sw_dedup.groupby("month").agg(gross=("gross", "sum"))
            .reset_index().assign(store="Tampa")
            .sort_values("month")
        )
        tampa_monthly["month_label"] = tampa_monthly["month"].dt.strftime("%b %Y")

        # All stores' May 2026 from per-store files (incl. Tampa's May, the only place
        # it now appears). This is the only period available for the non-Tampa stores.
        other_monthly = (
            store_df.groupby("store").agg(gross=("gross", "sum")).reset_index()
        )
        other_monthly["store"] = other_monthly["store"].str.replace("Eden - ", "", regex=False)
        other_monthly["month_label"] = "May 2026"

        # Chronological x-axis ordering across all months present (Mar–Jun 2026)
        _month_keys = list(tampa_monthly["month"]) + [pd.Timestamp("2026-05-01")]
        month_order = [pd.Timestamp(m).strftime("%b %Y") for m in sorted(set(_month_keys))]

        all_monthly = pd.concat([
            tampa_monthly[["month_label", "store", "gross"]],
            other_monthly[["month_label", "store", "gross"]],
        ])

        STORE_COLORS = {
            "Tampa":        "#2d6a4f",
            "Sarasota":     "#2563eb",
            "Cocoa Beach":  "#7c3aed",
            "Orlando":      "#d97706",
            "Delivery Hub": "#059669",
        }
        fig7 = go.Figure()
        for store in ["Tampa", "Sarasota", "Cocoa Beach", "Orlando", "Delivery Hub"]:
            d = all_monthly[all_monthly["store"] == store]
            if d.empty:
                continue
            fig7.add_trace(go.Bar(
                name=store, x=d["month_label"], y=d["gross"],
                marker_color=STORE_COLORS.get(store, "#888"),
                hovertemplate="%{x}  $%{y:,.0f}<extra>" + store + "</extra>",
            ))
        fig7.update_layout(
            barmode="stack", yaxis_title="",
            legend=dict(orientation="h", y=1.1),
            xaxis=dict(categoryorder="array", categoryarray=month_order),
        )
        chart_layout(fig7, "Monthly Gross Sales by Store", height=300)
        st.plotly_chart(fig7, use_container_width=True)

    with col_pie:
        pie_data      = cat_total.sort_values("gross", ascending=False)
        all_gross_pie = pie_data["gross"].sum()
        fig8 = go.Figure(go.Pie(
            labels=pie_data["category"], values=pie_data["gross"], hole=0.55,
            marker=dict(colors=[COLORS.get(c,"#888") for c in pie_data["category"]]),
            textinfo="label+percent", textfont=dict(color="#374151", size=11),
            hovertemplate="%{label}  $%{value:,.0f}<extra></extra>",
        ))
        fig8.add_annotation(text=f"${all_gross_pie:,.0f}", x=0.5, y=0.5,
                            showarrow=False, font=dict(color="#111827", size=13))
        chart_layout(fig8, "All-Store Sales Mix", height=300)
        st.plotly_chart(fig8, use_container_width=True)

    # ══════════════════════════════════════════════════════════════════════════
    # HEADSET ANALYTICS — weekly trend, delivery split, day/hour, demographics,
    # top products.  (Source: headset_reports_july_19, refreshed Jul 19 2026.)
    # ══════════════════════════════════════════════════════════════════════════
    if _headset_error:
        st.warning(f"Headset analytics unavailable: {_headset_error}")
    elif HS:
        st.divider()
        st.subheader("Headset Analytics")
        st.caption("Unified all-store reporting from Headset · through Jul 6, 2026")

        STORE_COLORS_HS = {
            "Tampa": "#2d6a4f", "Sarasota": "#2563eb",
            "Cocoa Beach": "#7c3aed", "Orlando": "#d97706",
        }

        # ── Weekly revenue trend by store ──────────────────────────────────────
        wk = HS["weekly"]
        if not wk.empty:
            fig_wk = go.Figure()
            for store in ["Tampa", "Sarasota", "Cocoa Beach", "Orlando"]:
                d = wk[wk["store"] == store].sort_values("week")
                if d.empty:
                    continue
                fig_wk.add_trace(go.Scatter(
                    name=store, x=d["week"], y=d["gross"], mode="lines+markers",
                    line=dict(color=STORE_COLORS_HS.get(store, "#888"), width=2.5),
                    marker=dict(size=5),
                    hovertemplate="%{x|%b %-d}  $%{y:,.0f}<extra>" + store + "</extra>",
                ))
            fig_wk.update_layout(legend=dict(orientation="h", y=1.12), yaxis_title="")
            chart_layout(fig_wk, "Weekly Gross Revenue by Store", height=340)
            st.plotly_chart(fig_wk, use_container_width=True)

        # ── In-store vs Delivery ───────────────────────────────────────────────
        pm = HS["pm_summary"]
        if not pm.empty:
            pm_map = {r["flag"]: r for _, r in pm.iterrows()}
            deliv = pm_map.get("Delivery", {})
            instore = pm_map.get("Not Delivery", {})
            d1, d2, d3, d4 = st.columns(4)
            for col, label, value, sub in [
                (d1, "In-Store Revenue",  f"${instore.get('gross',0):,.0f}",
                     f"${instore.get('avg_basket',0):,.0f} avg basket · {instore.get('units_per_basket',0):.1f} units"),
                (d2, "Delivery Revenue",  f"${deliv.get('gross',0):,.0f}",
                     f"${deliv.get('avg_basket',0):,.0f} avg basket · {deliv.get('units_per_basket',0):.1f} units"),
                (d3, "Delivery Basket vs In-Store",
                     f"{(deliv.get('avg_basket',0)/instore.get('avg_basket',1)):.1f}×" if instore.get('avg_basket',0) else "–",
                     "Larger delivery orders"),
                (d4, "Discount Rate",
                     f"{instore.get('discount_pct',0):.0f}% / {deliv.get('discount_pct',0):.0f}%",
                     "In-store / delivery"),
            ]:
                with col:
                    st.markdown(f"""<div class="kpi-card">
                      <div class="kpi-label">{label}</div>
                      <div class="kpi-value" style="font-size:1.4rem;">{value}</div>
                      <div class="kpi-sub">{sub}</div>
                    </div>""", unsafe_allow_html=True)
            st.markdown("<div style='height:14px;'></div>", unsafe_allow_html=True)

        col_pm, col_demo = st.columns(2)
        with col_pm:
            pmw = HS["pm_weekly"]
            if not pmw.empty:
                fig_pm = go.Figure()
                fig_pm.add_trace(go.Bar(
                    name="In-Store", x=pmw["week"], y=pmw["In-Store"], marker_color="#1a4731",
                    hovertemplate="%{x|%b %-d}  $%{y:,.0f}<extra>In-Store</extra>"))
                fig_pm.add_trace(go.Bar(
                    name="Delivery", x=pmw["week"], y=pmw["Delivery"], marker_color="#059669",
                    hovertemplate="%{x|%b %-d}  $%{y:,.0f}<extra>Delivery</extra>"))
                fig_pm.update_layout(barmode="stack", legend=dict(orientation="h", y=1.15),
                                     yaxis_title="")
                chart_layout(fig_pm, "Weekly Revenue — In-Store vs Delivery", height=320)
                st.plotly_chart(fig_pm, use_container_width=True)

        with col_demo:
            demo = HS["demographics"]
            if not demo.empty:
                age_order = ["25 & under", "26-35", "36-45", "46-55", "56-65", "66 & up"]
                demo = demo[demo["age"].isin(age_order)]
                fig_dm = go.Figure()
                for gender, color in [("M", "#2563eb"), ("F", "#db2777")]:
                    d = demo[demo["gender"] == gender].set_index("age").reindex(age_order).reset_index()
                    fig_dm.add_trace(go.Bar(
                        name={"M": "Male", "F": "Female"}[gender],
                        x=d["age"], y=d["gross"].fillna(0), marker_color=color,
                        hovertemplate="%{x}  $%{y:,.0f}<extra>" + gender + "</extra>"))
                fig_dm.update_layout(barmode="stack", legend=dict(orientation="h", y=1.15),
                                     yaxis_title="")
                chart_layout(fig_dm, "Revenue by Age Group & Gender", height=320)
                st.plotly_chart(fig_dm, use_container_width=True)

        # ── Sales by weekday × hour heatmap ────────────────────────────────────
        wh = HS["weekday_hour"]
        if not wh.empty:
            day_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
            days = [d for d in day_order if d in wh["day"].unique()]
            hours = sorted(wh["hour"].unique())
            grid = wh.pivot(index="day", columns="hour", values="gross").reindex(days)
            grid = grid.reindex(columns=hours)
            fig_hm = go.Figure(go.Heatmap(
                z=grid.values, x=[f"{h}:00" for h in hours], y=days,
                colorscale=[[0, "#f0f6f0"], [0.5, "#74c69d"], [1, "#1a4731"]],
                hovertemplate="%{y} %{x}  $%{z:,.0f}<extra></extra>",
                colorbar=dict(title="", thickness=10),
            ))
            chart_layout(fig_hm, "Sales by Day of Week & Hour", height=300)
            fig_hm.update_layout(yaxis=dict(autorange="reversed"))
            st.plotly_chart(fig_hm, use_container_width=True)

        # ── Top products table ─────────────────────────────────────────────────
        tp = HS["top_products"]
        if not tp.empty:
            st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)
            top15 = tp.sort_values("gross", ascending=False).head(15)
            rows_html = ""
            for _, r in top15.iterrows():
                gm = f"{r['gm_pct']:.0f}%" if pd.notna(r["gm_pct"]) else "–"
                cat_color = COLORS.get(r["category"], "#6b7280")
                rows_html += (
                    f"<tr><td>{r['product']}</td>"
                    f"<td><span style='color:{cat_color};font-weight:600;'>{r['category']}</span></td>"
                    f"<td style='text-align:right;'>${r['gross']:,.0f}</td>"
                    f"<td style='text-align:right;'>{int(r['units']):,}</td>"
                    f"<td style='text-align:right;'>${r['avg_price']:,.2f}</td>"
                    f"<td style='text-align:right;'>{gm}</td></tr>"
                )
            st.markdown(f"""<div class="data-card" style="height:auto;">
              <table class="data-table">
                <thead><tr>
                  <th>Product</th><th>Category</th>
                  <th style="text-align:right;">Revenue</th>
                  <th style="text-align:right;">Units</th>
                  <th style="text-align:right;">Avg Price</th>
                  <th style="text-align:right;">Margin</th>
                </tr></thead>
                <tbody>{rows_html}</tbody>
              </table>
              <p class="data-note">Top 15 products by gross revenue · Headset, all stores</p>
            </div>""", unsafe_allow_html=True)


if view == "Inventory":

    # ── Filters ───────────────────────────────────────────────────────────────
    # Strain exists only on the Biotrack pipeline (finished goods carry no strain),
    # so the Strain filter narrows the Production Pipeline charts only.
    _icat    = _options(disp_df["category"], vault_df["category"], bio_df["broad_cat"], store_df["category"])
    _istore  = _options(disp_df["org"].map(DISP_ORGS), store_df["store"].map(_clean_store))
    _istrain = _options(bio_df["strain"])
    st.markdown('<div class="filter-bar">Filters</div>', unsafe_allow_html=True)
    fi1, fi2, fi3 = st.columns(3)
    isel_cat    = fi1.multiselect("Category", _icat, default=[],
                                  placeholder="All categories", key="inv_cat") or _icat
    isel_store  = fi2.multiselect("Store", _istore, default=[],
                                  placeholder="All stores", key="inv_store") or _istore
    isel_strain = fi3.multiselect("Strain (pipeline)", _istrain, default=[],
                                  placeholder="All strains", key="inv_strain") or _istrain

    disp_df = disp_df[
        disp_df["category"].isin(isel_cat) &
        disp_df["org"].map(DISP_ORGS).isin(isel_store)
    ].copy()
    vault_df = vault_df[vault_df["category"].isin(isel_cat)].copy()       # central vault: no store dim
    store_df = store_df[
        store_df["category"].isin(isel_cat) &
        store_df["store"].map(_clean_store).isin(isel_store)
    ].copy()
    sw_df = sw_df[
        sw_df["category"].isin(isel_cat) &
        sw_df["store"].map(_clean_store).isin(isel_store)
    ].copy()
    bio_df = bio_df[
        bio_df["broad_cat"].isin(isel_cat) &
        bio_df["strain"].isin(isel_strain)
    ].copy()

    # Recompute KPI scalars from the filtered frames
    total_finished = int(disp_df["qty"].sum() + vault_df["qty"].sum())
    _pb = bio_df[bio_df["stage"].notna() & ~bio_df["stage"].str.contains("Approved|Sample", na=False)]
    flower_lbs = _pb[_pb["unit"] == "lbs"]["display_count"].sum()
    mfg_L      = _pb[_pb["unit"] == "L"]["display_count"].sum()
    _inv_store_sub = ", ".join(isel_store) if len(isel_store) <= 3 else f"{len(isel_store)} stores"

    # ── Inventory KPIs ────────────────────────────────────────────────────────
    disp_units  = disp_df["qty"].sum()
    vault_units = vault_df["qty"].sum()
    ik1, ik2, ik3, ik4, ik5 = st.columns(5)
    with ik1:
        st.markdown(f"""<div class="kpi-hero">
          <div class="kpi-label-hero">On-Hand Units</div>
          <div class="kpi-value-hero">{total_finished:,}</div>
          <div class="kpi-sub-hero">Dispensaries + hub</div>
        </div>""", unsafe_allow_html=True)
    for col, label, value, sub in [
        (ik2, "Dispensary Floor", f"{int(disp_units):,}",   _inv_store_sub),
        (ik3, "Processing Vault", f"{int(vault_units):,}",  "Finished · pre-distribution"),
        (ik4, "Bulk Flower",      f"{flower_lbs:,.0f} lbs", "In production pipeline"),
        (ik5, "Bulk Concentrate", f"{mfg_L:.1f} L",         "In production pipeline"),
    ]:
        with col:
            st.markdown(f"""<div class="kpi-card">
              <div class="kpi-label">{label}</div>
              <div class="kpi-value">{value}</div>
              <div class="kpi-sub">{sub}</div>
            </div>""", unsafe_allow_html=True)
    st.divider()

    # ── Pipeline ──────────────────────────────────────────────────────────────
    st.subheader("Production Pipeline")

    pipe_df = bio_df[bio_df["stage"].notna()].copy()

    def _stage_bar(df, unit_label, title, height=320):
        agg = df.groupby(["stage", "category"])["display_count"].sum().reset_index()
        agg["order"] = agg["stage"].map({s: i for i, s in enumerate(STAGE_ORDER)}).fillna(99)
        agg = agg.sort_values("order")
        fig = go.Figure()
        for cat in agg["category"].unique():
            d = agg[agg["category"] == cat]
            fig.add_trace(go.Bar(
                name=cat, x=d["stage"], y=d["display_count"],
                marker_color=COLORS.get(BIO_CAT_MAP.get(cat, "Other"), "#888"),
                hovertemplate=f"%{{x}}<br>%{{y:,.2f}} {unit_label}<extra>{cat}</extra>",
            ))
        fig.update_layout(barmode="stack", xaxis_tickangle=-30, yaxis_title=unit_label,
                          legend=dict(orientation="h", y=1.1, font=dict(size=10)))
        return chart_layout(fig, title, height=height)

    col_fl, col_mfg = st.columns(2)
    with col_fl:
        flower_pipe = pipe_df[pipe_df["unit"] == "lbs"]
        st.plotly_chart(
            _stage_bar(flower_pipe, "lbs", f"Flower  ·  {flower_pipe['display_count'].sum():,.0f} lbs total"),
            use_container_width=True)
    with col_mfg:
        mfg_pipe = pipe_df[pipe_df["unit"] == "L"]
        st.plotly_chart(
            _stage_bar(mfg_pipe, "L", f"Manufactured  ·  {mfg_pipe['display_count'].sum():,.2f} L total"),
            use_container_width=True)

    st.divider()

    # ── Inventory ─────────────────────────────────────────────────────────────
    st.subheader("Finished Goods Inventory")

    col_disp, col_vault = st.columns([3, 2])
    with col_disp:
        disp_agg = disp_df.copy()
        disp_agg["store"] = disp_agg["org"].map(DISP_ORGS)
        disp_agg = disp_agg.groupby(["store", "category"])["qty"].sum().reset_index()
        store_order = ["Tampa", "Sarasota", "Cocoa Beach", "Orlando"]
        fig3 = go.Figure()
        for cat in ["Flower","Concentrates","Edibles","Carts","Pre-rolls","Tinctures","Topical"]:
            d = disp_agg[disp_agg["category"] == cat]
            fig3.add_trace(go.Bar(
                name=cat, x=d["store"], y=d["qty"],
                marker_color=COLORS.get(cat, "#888"),
                hovertemplate="%{x} — " + cat + "<br>%{y:,} units<extra></extra>",
            ))
        fig3.update_layout(barmode="stack", yaxis_title="Units",
                           legend=dict(orientation="h", y=1.1),
                           xaxis=dict(categoryorder="array", categoryarray=store_order))
        chart_layout(fig3, "Dispensaries", height=320)
        st.plotly_chart(fig3, use_container_width=True)

    with col_vault:
        vault_agg   = vault_df.groupby("category")["qty"].sum().reset_index().sort_values("qty", ascending=False)
        vault_total = vault_agg["qty"].sum()
        fig4 = go.Figure(go.Pie(
            labels=vault_agg["category"], values=vault_agg["qty"], hole=0.55,
            marker=dict(colors=[COLORS.get(c, "#888") for c in vault_agg["category"]]),
            textinfo="label+percent", textfont=dict(color="#374151", size=11),
            hovertemplate="%{label}<br>%{value:,} units<extra></extra>",
        ))
        fig4.add_annotation(text=f"{vault_total:,}<br>units", x=0.5, y=0.5,
                            showarrow=False, font=dict(color="#111827", size=13))
        chart_layout(fig4, f"Processing Vault  ·  {vault_total:,} units", height=320)
        st.plotly_chart(fig4, use_container_width=True)

    st.divider()

    # ── Supply vs Demand ──────────────────────────────────────────────────────
    st.subheader("Weeks of Stock")

    floor_df = disp_df.groupby("category")["qty"].sum().reset_index()
    floor_df.columns = ["category", "on_hand"]

    # All-store weekly velocity by category:
    #   Tampa  → true last-4-week run-rate from daily Sweed data
    #   Others → May per-store snapshot (only period available) ÷ weeks per month
    WEEKS_PER_MONTH = 4.345
    cutoff   = sw_df["date"].max() - pd.Timedelta(weeks=4)
    tampa_v  = sw_df[sw_df["date"] >= cutoff].groupby("category")["qty"].sum() / 4
    other_v  = (
        store_df[store_df["store"] != "Eden - Tampa"]
        .groupby("category")["qty"].sum() / WEEKS_PER_MONTH
    )
    velocity = tampa_v.add(other_v, fill_value=0).rename("weekly_velocity").reset_index()
    svd = floor_df.merge(velocity, on="category", how="outer").fillna(0)
    svd["weeks_of_stock"] = svd.apply(
        lambda r: round(r["on_hand"] / r["weekly_velocity"], 1) if r["weekly_velocity"] > 0 else None,
        axis=1)

    col_svd, col_svd_tbl = st.columns([3, 2])
    with col_svd:
        fig9 = make_subplots(specs=[[{"secondary_y": True}]])
        fig9.add_trace(go.Bar(
            name="Units On-Hand", x=svd["category"], y=svd["on_hand"],
            marker_color=[COLORS.get(c, "#888") for c in svd["category"]], opacity=0.8,
            hovertemplate="%{x}  %{y:,} units<extra></extra>",
        ), secondary_y=False)
        fig9.add_trace(go.Scatter(
            name="Weekly Velocity (all stores)", x=svd["category"], y=svd["weekly_velocity"],
            mode="markers+lines", marker=dict(size=9, color="#d97706", symbol="diamond"),
            line=dict(color="#d97706", width=2, dash="dot"),
            hovertemplate="%{x}  ~%{y:.0f}/wk<extra></extra>",
        ), secondary_y=True)
        fig9.update_yaxes(title_text="Units", secondary_y=False,
                          gridcolor=GRID_COLOR, tickfont=dict(color=TEXT_COLOR))
        fig9.update_yaxes(title_text="Units/Week", secondary_y=True,
                          gridcolor="rgba(0,0,0,0)", tickfont=dict(color="#d97706"))
        fig9.update_layout(
            paper_bgcolor=CHART_BG, plot_bgcolor=CHART_BG, font=dict(color=TEXT_COLOR),
            height=300, margin=dict(l=10, r=10, t=40, b=10),
            legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color="#374151", size=11)),
            title=dict(text="Dispensary Stock vs. Weekly Sales Velocity",
                       font=dict(color="#111827", size=13), x=0.01),
            xaxis=dict(gridcolor=GRID_COLOR, tickfont=dict(color=TEXT_COLOR)),
        )
        st.plotly_chart(fig9, use_container_width=True)

    with col_svd_tbl:
        svd_d = svd[["category","on_hand","weekly_velocity","weeks_of_stock"]].copy()
        svd_d["on_hand_fmt"]   = svd_d["on_hand"].apply(lambda x: f"{int(x):,}")
        svd_d["velocity_fmt"]  = svd_d["weekly_velocity"].apply(lambda x: f"~{x:.0f}/wk" if pd.notna(x) and x > 0 else "–")
        svd_d["weeks_fmt"]     = svd_d["weeks_of_stock"].apply(lambda x: f"{x:.1f}" if pd.notna(x) else "–")
        rows_html = "".join(
            f"<tr><td>{r['category']}</td><td>{r['on_hand_fmt']}</td>"
            f"<td>{r['velocity_fmt']}</td><td>{r['weeks_fmt']}</td></tr>"
            for _, r in svd_d.sort_values("category").iterrows()
        )
        st.markdown(f"""<div class="data-card">
          <table class="data-table">
            <thead><tr><th>Category</th><th>On-Hand</th><th>Wkly Vel.</th><th>Wks Stock</th></tr></thead>
            <tbody>{rows_html}</tbody>
          </table>
          <p class="data-note">All-store weekly velocity &nbsp;·&nbsp; +{vault_df['qty'].sum():,} units in processing vault</p>
        </div>""", unsafe_allow_html=True)

    st.divider()

    # ── Weeks on Hand by Location ─────────────────────────────────────────────
    st.subheader("Weeks on Hand by Location")

    # Finished on-hand units per dispensary (Tampa, Sarasota, Cocoa Beach, Orlando)
    loc_oh = disp_df.copy()
    loc_oh["store"] = loc_oh["org"].map(DISP_ORGS)
    loc_oh = loc_oh.groupby("store")["qty"].sum().reset_index()
    loc_oh.columns = ["store", "on_hand"]

    # Weekly sales velocity per location:
    #   Tampa  → true last-4-week run-rate from daily Sweed data.
    #   Others → May per-store snapshot (only period available) ÷ weeks per month.
    WEEKS_PER_MONTH = 4.345
    tampa_vel = (
        sw_df[sw_df["date"] >= sw_df["date"].max() - pd.Timedelta(weeks=4)]["qty"].sum() / 4
    )
    store_vel = store_df.groupby("store")["qty"].sum().reset_index()
    store_vel["store"] = store_vel["store"].str.replace("Eden - ", "", regex=False)
    vel_map = {s: q / WEEKS_PER_MONTH for s, q in zip(store_vel["store"], store_vel["qty"])}
    vel_map["Tampa"] = tampa_vel  # daily-derived run-rate beats the 1-month snapshot

    woh = loc_oh.copy()
    woh["weekly_velocity"] = woh["store"].map(vel_map).fillna(0)
    woh["weeks_on_hand"] = woh.apply(
        lambda r: round(r["on_hand"] / r["weekly_velocity"], 1) if r["weekly_velocity"] > 0 else None,
        axis=1)
    woh = woh.sort_values("weeks_on_hand", ascending=True, na_position="last")

    def _woh_color(w):
        if w is None or pd.isna(w): return "#9ca3af"  # no velocity → gray
        if w < 2:  return "#dc2626"   # critical  (<2 wks)
        if w < 4:  return "#d97706"   # watch     (2–4 wks)
        if w <= 8: return "#16a34a"   # healthy   (4–8 wks)
        return "#2563eb"              # overstock (>8 wks)

    col_woh, col_woh_tbl = st.columns([3, 2])
    with col_woh:
        plot = woh[woh["weeks_on_hand"].notna()]
        fig10 = go.Figure(go.Bar(
            x=plot["weeks_on_hand"], y=plot["store"], orientation="h",
            marker_color=[_woh_color(w) for w in plot["weeks_on_hand"]],
            text=[f"{w:.1f} wks" for w in plot["weeks_on_hand"]],
            textposition="outside", cliponaxis=False,
            hovertemplate="%{y}<br>%{x:.1f} weeks on hand<extra></extra>",
        ))
        fig10.update_layout(
            paper_bgcolor=CHART_BG, plot_bgcolor=CHART_BG, font=dict(color=TEXT_COLOR),
            height=300, margin=dict(l=10, r=50, t=40, b=10),
            title=dict(text="Finished Inventory Coverage (weeks at current run-rate)",
                       font=dict(color="#111827", size=13), x=0.01),
            xaxis=dict(title="Weeks on hand", gridcolor=GRID_COLOR, tickfont=dict(color=TEXT_COLOR)),
            yaxis=dict(gridcolor="rgba(0,0,0,0)", tickfont=dict(color=TEXT_COLOR)),
        )
        st.plotly_chart(fig10, use_container_width=True)

    with col_woh_tbl:
        wd = woh.copy()
        wd["on_hand_fmt"] = wd["on_hand"].apply(lambda x: f"{int(x):,}")
        wd["vel_fmt"]     = wd["weekly_velocity"].apply(lambda x: f"~{x:.0f}/wk" if pd.notna(x) and x > 0 else "–")
        wd["woh_fmt"]     = wd["weeks_on_hand"].apply(lambda x: f"{x:.1f}" if pd.notna(x) else "–")
        rows_html = "".join(
            f"<tr><td>{r['store']}</td><td>{r['on_hand_fmt']}</td>"
            f"<td>{r['vel_fmt']}</td><td>{r['woh_fmt']}</td></tr>"
            for _, r in wd.iterrows()
        )
        st.markdown(f"""<div class="data-card">
          <table class="data-table">
            <thead><tr><th>Location</th><th>On-Hand</th><th>Wkly Vel.</th><th>Wks on Hand</th></tr></thead>
            <tbody>{rows_html}</tbody>
          </table>
          <p class="data-note">All-store weekly sales velocity</p>
        </div>""", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════════
    # HEADSET SUPPLY ANALYTICS — days-of-supply, dead stock, reorder watchlist.
    # (Source: headset_reports_july_19, refreshed Jul 19 2026.)
    # ══════════════════════════════════════════════════════════════════════════
    if _headset_error:
        st.warning(f"Headset analytics unavailable: {_headset_error}")
    elif HS:
        st.divider()
        st.subheader("Headset Supply Analytics")
        st.caption("Unified all-store inventory from Headset · refreshed Jul 19, 2026")

        col_ds, col_dead = st.columns(2)
        with col_ds:
            ds = HS["days_supply"].copy()
            if not ds.empty:
                ds = ds.sort_values("days_remaining", ascending=True)

                def _ds_color(d):
                    if d < 30:   return "#dc2626"   # critical
                    if d < 90:   return "#d97706"   # watch
                    if d <= 365: return "#16a34a"   # healthy
                    return "#2563eb"                # overstock

                fig_ds = go.Figure(go.Bar(
                    x=ds["days_remaining"], y=ds["category"], orientation="h",
                    marker_color=[_ds_color(d) for d in ds["days_remaining"]],
                    text=[f"{d:,.0f} d" for d in ds["days_remaining"]],
                    textposition="outside", cliponaxis=False,
                    hovertemplate="%{y}<br>%{x:,.0f} days · %{customdata:,.0f} units on hand<extra></extra>",
                    customdata=ds["on_hand"],
                ))
                chart_layout(fig_ds, "Est. Days of Supply by Category", height=320)
                fig_ds.update_layout(
                    xaxis=dict(title="Days remaining", range=[0, ds["days_remaining"].max() * 1.18]),
                    yaxis=dict(gridcolor="rgba(0,0,0,0)"),
                    margin=dict(l=10, r=60, t=40, b=10),
                )
                st.plotly_chart(fig_ds, use_container_width=True)

        with col_dead:
            dead = HS["dead_stock"].copy()
            if not dead.empty:
                dead = dead.sort_values("skus", ascending=True)
                fig_dead = go.Figure(go.Bar(
                    x=dead["skus"], y=dead["category"], orientation="h",
                    marker_color=[COLORS.get(c, "#9ca3af") for c in dead["category"]],
                    text=dead["skus"], textposition="outside", cliponaxis=False,
                    hovertemplate="%{y}<br>%{x} SKUs in stock with no sales<extra></extra>",
                ))
                chart_layout(fig_dead, "Dead Stock — In-Stock SKUs with No Sales", height=320)
                fig_dead.update_layout(
                    xaxis=dict(title="Distinct SKUs", range=[0, dead["skus"].max() * 1.18]),
                    yaxis=dict(gridcolor="rgba(0,0,0,0)"),
                    margin=dict(l=10, r=50, t=40, b=10),
                )
                st.plotly_chart(fig_dead, use_container_width=True)

        # ── Reorder watchlist — lowest days-remaining SKUs actively selling ─────
        pl = HS["product_list"].copy()
        pl = pl[pl["store"].map(_clean_store).isin(isel_store) & pl["category"].isin(isel_cat)]
        watch = pl[(pl["days_remaining"].notna()) & (pl["per_day"] > 0)].copy()
        watch = watch.sort_values("days_remaining", ascending=True).head(15)
        if not watch.empty:
            st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)
            rows_html = ""
            for _, r in watch.iterrows():
                d = r["days_remaining"]
                color = "#dc2626" if d < 7 else ("#d97706" if d < 21 else "#374151")
                sug = f"{int(r['suggested_order'])}" if r["suggested_order"] > 0 else "–"
                rows_html += (
                    f"<tr><td>{r['product']}</td><td>{r['store']}</td>"
                    f"<td>{r['category']}</td>"
                    f"<td style='text-align:right;'>{int(r['on_hand']):,}</td>"
                    f"<td style='text-align:right;'>{r['per_day']:.1f}</td>"
                    f"<td style='text-align:right;color:{color};font-weight:600;'>{d:,.0f}</td>"
                    f"<td style='text-align:right;'>{sug}</td></tr>"
                )
            st.markdown(f"""<div class="data-card" style="height:auto;">
              <table class="data-table">
                <thead><tr>
                  <th>Product</th><th>Store</th><th>Category</th>
                  <th style="text-align:right;">On-Hand</th>
                  <th style="text-align:right;">Units/Day</th>
                  <th style="text-align:right;">Days Left</th>
                  <th style="text-align:right;">Sug. Order</th>
                </tr></thead>
                <tbody>{rows_html}</tbody>
              </table>
              <p class="data-note">15 fastest-to-stockout SKUs (currently selling) · Headset</p>
            </div>""", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════════
    # MONDAY OPS BOARDS — production supplies reorder + finished-goods batches.
    # (Source: monday_reports_july_19, exported Jul 16–19 2026.)
    # ══════════════════════════════════════════════════════════════════════════
    if _monday_inv_error:
        st.warning(f"Monday ops boards unavailable: {_monday_inv_error}")
    elif MINV:
        # ── Production supplies reorder ────────────────────────────────────────
        sup = MINV.get("supplies", pd.DataFrame())
        if not sup.empty:
            st.divider()
            st.subheader("Production Supplies")
            st.caption("Packaging & input inventory vs. reorder levels · Monday Inventory Management board")

            n_emergency = int((sup["status"] == "Emergency Order Needed").sum())
            n_reorder   = int((sup["status"] == "Reorder Needed").sum())
            n_ok        = int((sup["status"] == "Sufficient").sum())
            s1, s2, s3 = st.columns(3)
            for col, label, value, sub, bg in [
                (s1, "Emergency Orders", n_emergency, "Below critical level", "#fee2e2"),
                (s2, "Reorder Needed",   n_reorder,   "Below reorder level",  "#fef3c7"),
                (s3, "Sufficient",       n_ok,        "At or above target",   "#d1fae5"),
            ]:
                with col:
                    st.markdown(f"""<div class="kpi-card" style="background:{bg};">
                      <div class="kpi-label">{label}</div>
                      <div class="kpi-value">{value}</div>
                      <div class="kpi-sub">{sub}</div>
                    </div>""", unsafe_allow_html=True)
            st.markdown("<div style='height:14px;'></div>", unsafe_allow_html=True)

            rank = {"Emergency Order Needed": 0, "Reorder Needed": 1, "Sufficient": 2}
            need = sup[sup["status"] != "Sufficient"].copy()
            need = need.sort_values(by="status", key=lambda s: s.map(rank))
            rows_html = ""
            for _, r in need.iterrows():
                is_emerg = r["status"] == "Emergency Order Needed"
                badge_bg, badge_fg = ("#fee2e2", "#b91c1c") if is_emerg else ("#fef3c7", "#b45309")
                label = "Emergency" if is_emerg else "Reorder"
                oh = f"{int(r['on_hand']):,}" if pd.notna(r["on_hand"]) else "–"
                rl = f"{int(r['reorder_level']):,}" if pd.notna(r["reorder_level"]) else "–"
                rows_html += (
                    f"<tr><td>{r['item']}</td>"
                    f"<td><span style='background:{badge_bg};color:{badge_fg};padding:2px 8px;"
                    f"border-radius:6px;font-size:0.72rem;font-weight:700;'>{label}</span></td>"
                    f"<td style='text-align:right;'>{oh}</td>"
                    f"<td style='text-align:right;'>{rl}</td>"
                    f"<td>{r['vendor'] or '–'}</td></tr>"
                )
            st.markdown(f"""<div class="data-card" style="height:auto;">
              <table class="data-table">
                <thead><tr>
                  <th>Supply Item</th><th>Status</th>
                  <th style="text-align:right;">On Hand</th>
                  <th style="text-align:right;">Reorder Level</th>
                  <th>Vendor</th>
                </tr></thead>
                <tbody>{rows_html}</tbody>
              </table>
              <p class="data-note">{n_emergency + n_reorder} items need ordering · placed Tuesdays weekly</p>
            </div>""", unsafe_allow_html=True)

        # ── Finished-goods batches & expiry ────────────────────────────────────
        fg = MINV.get("finished_goods", pd.DataFrame())
        if not fg.empty:
            st.divider()
            st.subheader("Finished Goods Batches")
            st.caption("Batch-level stock, testing & expiration · Monday Finished Goods Tracker")

            HIST = {"Expired Products", "Sold Out"}
            live = fg[~fg["group"].isin(HIST)].copy()

            col_fgcat, col_exp = st.columns([2, 3])
            with col_fgcat:
                by_cat = (live[live["on_hand"] > 0].groupby("group")["on_hand"].sum()
                          .reset_index().sort_values("on_hand", ascending=True))
                fig_fg = go.Figure(go.Bar(
                    x=by_cat["on_hand"], y=by_cat["group"], orientation="h",
                    marker_color="#2d6a4f",
                    text=by_cat["on_hand"].apply(lambda x: f"{int(x):,}"),
                    textposition="outside", cliponaxis=False,
                    hovertemplate="%{y}<br>%{x:,.0f} units on hand<extra></extra>",
                ))
                chart_layout(fig_fg, "On-Hand Units by Product Type", height=340)
                fig_fg.update_layout(
                    xaxis=dict(range=[0, by_cat["on_hand"].max() * 1.2]),
                    yaxis=dict(gridcolor="rgba(0,0,0,0)"),
                    margin=dict(l=10, r=60, t=40, b=10),
                )
                st.plotly_chart(fig_fg, use_container_width=True)

            with col_exp:
                soon = fg[(fg["days_to_exp"].notna()) & (fg["on_hand"] > 0)
                          & (fg["days_to_exp"] <= 90)].copy()
                soon = soon.sort_values("days_to_exp").head(12)
                if soon.empty:
                    st.markdown("""<div class="data-card" style="height:340px;justify-content:center;">
                      <p class="data-note" style="text-align:center;font-size:0.85rem;">
                      No in-stock batches expiring within 90 days ✓</p></div>""",
                                unsafe_allow_html=True)
                else:
                    rows_html = ""
                    for _, r in soon.iterrows():
                        d = int(r["days_to_exp"])
                        if d < 0:
                            badge = f"<span style='color:#b91c1c;font-weight:700;'>Expired {abs(d)}d</span>"
                        elif d <= 30:
                            badge = f"<span style='color:#b91c1c;font-weight:700;'>{d}d</span>"
                        else:
                            badge = f"<span style='color:#b45309;font-weight:600;'>{d}d</span>"
                        rows_html += (
                            f"<tr><td>{r['product']}</td><td>{r['group']}</td>"
                            f"<td style='text-align:right;'>{int(r['on_hand']):,}</td>"
                            f"<td style='text-align:right;'>{badge}</td></tr>"
                        )
                    st.markdown(f"""<div class="data-card" style="height:auto;min-height:340px;">
                      <table class="data-table">
                        <thead><tr>
                          <th>Batch</th><th>Type</th>
                          <th style="text-align:right;">On Hand</th>
                          <th style="text-align:right;">Expires</th>
                        </tr></thead>
                        <tbody>{rows_html}</tbody>
                      </table>
                      <p class="data-note">In-stock batches expiring within 90 days (negative = already expired)</p>
                    </div>""", unsafe_allow_html=True)


if view == "Harvest":

    if _harvest_error:
        st.error(f"Could not load harvest data: {_harvest_error}")

    # ── Filters ───────────────────────────────────────────────────────────────
    # Harvest data carries strain + harvest date (no category / store dimension).
    _hd = harvested["harvest_date"].dropna()
    _hmin = _hd.min().date() if not _hd.empty else None
    _hmax = _hd.max().date() if not _hd.empty else None
    _hstrain = _options(pd.concat([harvested["strain"], active["strain"], upcoming["strain"]]))
    st.markdown('<div class="filter-bar">Filters</div>', unsafe_allow_html=True)
    hf1, hf2 = st.columns([2, 3])
    if _hmin and _hmax and _hmin < _hmax:
        hsel_range = hf1.date_input("Harvest date range", (_hmin, _hmax),
                                    min_value=_hmin, max_value=_hmax, key="hv_date")
    else:
        hsel_range = None
    hsel_strain = hf2.multiselect("Strain", _hstrain, default=[],
                                  placeholder="All strains", key="hv_strain") or _hstrain

    if isinstance(hsel_range, (list, tuple)) and len(hsel_range) == 2:
        hd0, hd1 = pd.Timestamp(hsel_range[0]), pd.Timestamp(hsel_range[1]) + pd.Timedelta(days=1)
    else:
        hd0, hd1 = None, None

    # Strain filter narrows every harvest table; the date filter narrows actual
    # harvests (active/upcoming hold planned/future dates, so they stay strain-only).
    harvested = harvested[harvested["strain"].isin(hsel_strain)].copy()
    if hd0 is not None:
        harvested = harvested[
            (harvested["harvest_date"] >= hd0) & (harvested["harvest_date"] < hd1)
        ].copy()
    active   = active[active["strain"].isin(hsel_strain)].copy()
    upcoming = upcoming[upcoming["strain"].isin(hsel_strain)].copy()

    # ── KPIs ──────────────────────────────────────────────────────────────────
    total_cycles  = len(harvested)
    total_dry_lbs = harvested["dry_lbs"].sum()
    total_ff_lbs  = harvested["ff_lbs"].sum()
    total_wet_lbs = harvested["wet_lbs"].sum()
    avg_dry_ratio = (total_dry_lbs / total_wet_lbs * 100) if total_wet_lbs > 0 else 0
    avg_per_plant = (harvested["dry_weight_g"] / harvested["num_plants"]).dropna().mean() / G_PER_LB
    _hdates = harvested["harvest_date"].dropna()
    _hrange = f"{_hdates.min():%b %Y} – {_hdates.max():%b %Y}" if not _hdates.empty else "All cycles"

    h1, h2, h3, h4, h5 = st.columns(5)
    with h1:
        st.markdown(f"""<div class="kpi-hero">
          <div class="kpi-label-hero">Cycles Harvested</div>
          <div class="kpi-value-hero">{total_cycles}</div>
          <div class="kpi-sub-hero">{_hrange}</div>
        </div>""", unsafe_allow_html=True)
    for col, label, value, sub in [
        (h2, "Total Dry Yield",    f"{total_dry_lbs:,.0f} lbs", "All cycles"),
        (h3, "Total Fresh Frozen", f"{total_ff_lbs:,.0f} lbs",  "Extract feedstock"),
        (h4, "Wet → Dry Ratio",    f"{avg_dry_ratio:.1f}%",     "Avg across all harvests"),
        (h5, "Yield / Plant",      f"{avg_per_plant:.2f} lbs",  "Avg dry weight"),
    ]:
        with col:
            st.markdown(f"""<div class="kpi-card">
              <div class="kpi-label">{label}</div>
              <div class="kpi-value">{value}</div>
              <div class="kpi-sub">{sub}</div>
            </div>""", unsafe_allow_html=True)

    st.divider()

    # ── Monthly output ────────────────────────────────────────────────────────
    st.subheader("Monthly Harvest Output")

    monthly_h = harvested.dropna(subset=["month"]).groupby("month").agg(
        dry_lbs=("dry_lbs","sum"), ff_lbs=("ff_lbs","sum"),
        trim_lbs=("trim_lbs","sum"), cycles=("strain","count"),
    ).reset_index().sort_values("month")

    col_m1, col_m2 = st.columns([3, 2])
    with col_m1:
        fig_mo = go.Figure()
        for name, col_data, color in [
            ("Dry Flower", monthly_h["dry_lbs"],  "#2d6a4f"),
            ("Fresh Frozen", monthly_h["ff_lbs"], "#2563eb"),
            ("Trim",        monthly_h["trim_lbs"],"#d97706"),
        ]:
            fig_mo.add_trace(go.Bar(name=name, x=monthly_h["month"], y=col_data,
                                    marker_color=color,
                                    hovertemplate="%{x|%b %Y}  %{y:,.0f} lbs<extra>" + name + "</extra>"))
        fig_mo.add_trace(go.Scatter(
            name="Cycles", x=monthly_h["month"], y=monthly_h["cycles"],
            mode="markers+lines", yaxis="y2",
            marker=dict(size=7, color="#d97706"), line=dict(color="#d97706", width=2, dash="dot"),
            hovertemplate="%{x|%b %Y}  %{y} cycles<extra></extra>",
        ))
        fig_mo.update_layout(
            barmode="stack", yaxis_title="lbs",
            yaxis2=dict(title="Cycles", overlaying="y", side="right",
                        gridcolor="rgba(0,0,0,0)", tickfont=dict(color="#d97706")),
            legend=dict(orientation="h", y=1.1),
        )
        chart_layout(fig_mo, "Dry Flower, Fresh Frozen & Trim (lbs)", height=340)
        st.plotly_chart(fig_mo, use_container_width=True)

    with col_m2:
        des_agg = harvested.groupby("designation").agg(dry_lbs=("dry_lbs","sum")).reset_index()
        des_agg = des_agg[des_agg["designation"].notna() & (des_agg["designation"] != "")]
        des_colors = {"Flower":"#2d6a4f","Wash":"#2563eb","Pre-rolls":"#d97706","TBD":"#9ca3af"}
        fig_des = go.Figure(go.Pie(
            labels=des_agg["designation"], values=des_agg["dry_lbs"], hole=0.55,
            marker=dict(colors=[des_colors.get(d,"#888") for d in des_agg["designation"]]),
            textinfo="label+percent", textfont=dict(color=TEXT_COLOR, size=11),
            hovertemplate="%{label}  %{value:,.0f} lbs<extra></extra>",
        ))
        fig_des.add_annotation(text=f"{des_agg['dry_lbs'].sum():,.0f}<br>lbs",
                               x=0.5, y=0.5, showarrow=False, font=dict(color="#111827", size=12))
        chart_layout(fig_des, "Designation Mix", height=340)
        st.plotly_chart(fig_des, use_container_width=True)

    st.divider()

    # ── Room performance ──────────────────────────────────────────────────────
    st.subheader("Room Performance")

    room_agg = harvested.dropna(subset=["room"]).groupby("room").agg(
        dry_lbs=("dry_lbs","sum"), ff_lbs=("ff_lbs","sum"),
        cycles=("strain","count"), plants=("num_plants","sum"),
    ).reset_index()
    room_agg["dry_per_plant"] = (room_agg["dry_lbs"] / room_agg["plants"]).round(3)

    col_r1, col_r2 = st.columns(2)
    with col_r1:
        fig_room = go.Figure()
        fig_room.add_trace(go.Bar(name="Dry Flower", x=room_agg["room"], y=room_agg["dry_lbs"],
            marker_color=[ROOM_COLORS.get(r,"#888") for r in room_agg["room"]],
            hovertemplate="%{x}  %{y:,.0f} lbs<extra></extra>"))
        fig_room.add_trace(go.Bar(name="Fresh Frozen", x=room_agg["room"], y=room_agg["ff_lbs"],
            marker_color=[ROOM_COLORS.get(r,"#888") for r in room_agg["room"]],
            opacity=0.5,
            hovertemplate="%{x}  %{y:,.0f} lbs FF<extra></extra>"))
        fig_room.update_layout(barmode="group", yaxis_title="lbs", legend=dict(orientation="h", y=1.1))
        chart_layout(fig_room, "Output by Room", height=300)
        st.plotly_chart(fig_room, use_container_width=True)

    with col_r2:
        fig_eff = go.Figure(go.Bar(
            x=room_agg["room"], y=room_agg["dry_per_plant"],
            marker_color=[ROOM_COLORS.get(r,"#888") for r in room_agg["room"]],
            text=room_agg["dry_per_plant"].apply(lambda x: f"{x:.3f}"),
            textposition="outside", textfont=dict(color="#374151"),
            cliponaxis=False,
            hovertemplate="%{x}  %{y:.3f} lbs/plant<extra></extra>",
        ))
        chart_layout(fig_eff, "Avg Dry Yield per Plant (lbs)", height=300)
        fig_eff.update_layout(yaxis_title="lbs / plant")
        st.plotly_chart(fig_eff, use_container_width=True)

    st.divider()

    # ── Strains ───────────────────────────────────────────────────────────────
    st.subheader("Strain Performance")

    col_s1, col_s2 = st.columns([3, 2])
    with col_s1:
        strain_agg = harvested.groupby("strain").agg(
            dry_lbs=("dry_lbs","sum"), wet_lbs=("wet_lbs","sum"),
            cycles=("cycle","count"), plants=("num_plants","sum"),
        ).reset_index()
        strain_agg["dry_ratio"] = (strain_agg["dry_lbs"] / strain_agg["wet_lbs"] * 100).round(1)
        strain_top = strain_agg.sort_values("dry_lbs", ascending=True).tail(20)
        fig_st = go.Figure(go.Bar(
            name="Dry lbs", y=strain_top["strain"], x=strain_top["dry_lbs"],
            orientation="h", marker_color="#2d6a4f",
            hovertemplate="%{y}  %{x:,.0f} lbs (%{customdata:.1f}% of wet)<extra></extra>",
            customdata=strain_top["dry_ratio"],
        ))
        chart_layout(fig_st, "Top 20 Strains — Dry Yield (lbs)", height=460)
        fig_st.update_layout(xaxis_title="", yaxis=dict(gridcolor="rgba(0,0,0,0)"))
        st.plotly_chart(fig_st, use_container_width=True)

    with col_s2:
        thc_data = harvested.dropna(subset=["thc_pct"]).sort_values("thc_pct", ascending=True)
        if not thc_data.empty:
            fig_thc = go.Figure(go.Bar(
                y=thc_data["strain"], x=thc_data["thc_pct"], orientation="h",
                marker_color="#7c3aed",
                text=thc_data.apply(
                    lambda r: f"{r['thc_pct']:.1f}%  {r['terps_pct']:.2f}% terps"
                              if r["terps_pct"] else f"{r['thc_pct']:.1f}%", axis=1),
                textposition="outside", textfont=dict(color="#374151", size=11),
                cliponaxis=False,
                hovertemplate="%{y}  THC %{x:.1f}%<extra></extra>",
            ))
            chart_layout(fig_thc, "COA Results — THC %", height=240)
            fig_thc.update_layout(xaxis=dict(range=[0, max(thc_data["thc_pct"]) * 1.5]),
                                  yaxis=dict(gridcolor="rgba(0,0,0,0)"))
            st.plotly_chart(fig_thc, use_container_width=True)

        scatter_data = harvested.dropna(subset=["wet_lbs","dry_lbs"])
        scatter_data = scatter_data[scatter_data["wet_lbs"] > 0]
        fig_scat = go.Figure()
        for room in scatter_data["room"].dropna().unique():
            d = scatter_data[scatter_data["room"] == room]
            fig_scat.add_trace(go.Scatter(
                x=d["wet_lbs"], y=d["dry_lbs"], name=room, mode="markers",
                marker=dict(size=7, color=ROOM_COLORS.get(room,"#888"), opacity=0.75),
                hovertemplate="%{text}<br>Wet %{x:,.0f} · Dry %{y:,.0f} lbs<extra></extra>",
                text=d["strain"],
            ))
        chart_layout(fig_scat, "Wet vs. Dry per Cycle", height=240)
        fig_scat.update_layout(xaxis_title="Wet lbs", yaxis_title="Dry lbs",
                               legend=dict(orientation="h", y=1.12))
        st.plotly_chart(fig_scat, use_container_width=True)

    st.divider()

    # ── Active & upcoming ─────────────────────────────────────────────────────
    st.subheader("Active & Upcoming Cycles")

    def _html_table(df, note=None):
        cols = list(df.columns)
        header = "".join(f"<th>{c}</th>" for c in cols)
        body = "".join(
            "<tr>" + "".join(f"<td>{row[c] if pd.notna(row[c]) else '–'}</td>" for c in cols) + "</tr>"
            for _, row in df.iterrows()
        )
        note_html = f'<p class="data-note">{note}</p>' if note else ""
        return f"""<div class="data-card" style="height:auto;max-height:360px;overflow-y:auto;">
          <table class="data-table">
            <thead><tr>{header}</tr></thead>
            <tbody>{body}</tbody>
          </table>{note_html}
        </div>"""

    col_a1, col_a2 = st.columns(2)
    with col_a1:
        active_display = active[["strain","status","harvest_date","room","cycle"]].copy()
        active_display["harvest_date"] = active_display["harvest_date"].dt.strftime("%b %d, %Y").fillna("–")
        active_display.columns = ["Strain","Status","Est. Harvest","Room","Cycle"]
        st.markdown(_html_table(active_display.sort_values("Est. Harvest").fillna("–"),
                                note="Currently flowering / in trim"), unsafe_allow_html=True)

    with col_a2:
        upcoming_display = upcoming[["strain","harvest_date"]].copy()
        upcoming_display["harvest_date"] = upcoming_display["harvest_date"].dt.strftime("%b %d, %Y").fillna("–")
        upcoming_display.columns = ["Strain","Planned Harvest"]
        st.markdown(_html_table(upcoming_display.sort_values("Planned Harvest").fillna("–"),
                                note="Upcoming planned cycles"), unsafe_allow_html=True)

    st.divider()
    st.subheader("Full Harvest Log")
    log = harvested[["strain","harvest_date","room","cycle","designation",
                      "dry_lbs","ff_lbs","trim_lbs","thc_pct","terps_pct","num_plants"]].copy()
    log["harvest_date"] = log["harvest_date"].dt.strftime("%b %d, %Y")
    log["dry_lbs"]   = log["dry_lbs"].apply(lambda x: f"{x:,.2f}" if pd.notna(x) else "–")
    log["ff_lbs"]    = log["ff_lbs"].apply(lambda x: f"{x:,.2f}" if pd.notna(x) else "–")
    log["trim_lbs"]  = log["trim_lbs"].apply(lambda x: f"{x:,.2f}" if pd.notna(x) else "–")
    log["thc_pct"]   = log["thc_pct"].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "–")
    log["terps_pct"] = log["terps_pct"].apply(lambda x: f"{x:.3f}%" if pd.notna(x) else "–")
    log["num_plants"]= log["num_plants"].apply(lambda x: f"{int(x)}" if pd.notna(x) else "–")
    log.columns = ["Strain","Date","Room","Cycle","Designation",
                   "Dry (lbs)","FF (lbs)","Trim (lbs)","THC%","Terps%","Plants"]
    st.markdown(
        f"""<div class="data-card" style="height:auto;max-height:500px;overflow-y:auto;overflow-x:auto;">
          <table class="data-table">
            <thead><tr>{"".join(f"<th>{c}</th>" for c in log.columns)}</tr></thead>
            <tbody>{"".join(
              "<tr>" + "".join(f"<td>{row[c]}</td>" for c in log.columns) + "</tr>"
              for _, row in log.sort_values("Date", ascending=False).iterrows()
            )}</tbody>
          </table>
        </div>""", unsafe_allow_html=True)


st.caption("Eden Cannabis · Confidential · June 2026")
